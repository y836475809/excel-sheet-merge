# coding=utf-8
import csv
import os
import subprocess
import sys
from typing import List

import openpyxl

import em_util


def parse_patch(lines: List[str]):
    commands = []
    sheet_name = lines[0].split("/")[-1].replace(".csv", "")
    diffs = lines[5:]
    for i, d in enumerate(diffs):
        if d.startswith("@@ "):
            subcommands = []
            info = d.replace("@@ ", "").replace(" @@", "")
            row_index = parse_line_info(info)
            ddm = []
            ddp = []
            for x in range(i+1, len(diffs)):
                if diffs[x].startswith("-"):
                    ddm.append(diffs[x])
                elif diffs[x].startswith("+"):
                    ddp.append(diffs[x][1:])
                elif diffs[x].startswith("@@ "):
                    break
                else:
                    break
            df = len(ddp) - len(ddm)
            if df > 0:
                subcommands.append({
                    "cmd": "addrow",
                    "row": row_index,
                    "range": df,
                    "sheetname": sheet_name
                })
            elif df < 0:
                subcommands.append({
                    "cmd": "delrow",
                    "row": row_index,
                    "range": abs(df),
                    "sheetname": sheet_name
                })
            if len(ddp) > 0:
                subcommands.append({
                    "cmd": "setvalue",
                    "row": row_index,
                    "data": ddp,
                    "sheetname": sheet_name
                })
            commands.append(subcommands)
    return commands


def parse_line_info(line_info: str):
    diff = line_info.replace("-", "").replace("+", "").split(" ")
    pre = diff[0]
    post = diff[1]
    if pre.find(",") == -1:
        return int(pre)
    else:
        return int(pre.split(",")[0])


def sheet_merge(wb, cmds: List):
    cmds.reverse()
    print(f"merge: {cmds[0][0]['sheetname']}")
    for subcmds in cmds:
        sheetname = subcmds[0]["sheetname"]
        if sheetname in wb.sheetnames:
            ws = wb[sheetname]
        else:
            wb.create_sheet(index=0, title=sheetname)
            ws = wb[sheetname]
        row_offset = em_util.get_row_offset(ws)
        do_addrow = False
        for c in subcmds:
            row_index = c["row"] + row_offset
            if c["cmd"] == "addrow":
                print(f"\t{c['cmd']} row_index={row_index+1}, range={c['range']}")
                for ri in range(c["range"]):
                    do_addrow = True
                    ws.insert_rows(row_index + 1)
            if c["cmd"] == "delrow":
                print(f"\t{c['cmd']} row_index={row_index+1}, range={c['range']}")
                for ri in reversed(range(c["range"])):
                    ws.delete_rows(row_index + ri)
            if c["cmd"] == "setvalue":
                row_shift = 0
                if do_addrow:
                    row_shift = 1
                row = row_index
                csv_lines = list(csv.reader(
                    c["data"], quotechar='"', delimiter=',',
                    quoting=csv.QUOTE_MINIMAL, skipinitialspace=True))
                print(f"\tsetvalue row_index={row + row_shift}, range={len(csv_lines)}")
                for csv_values in csv_lines:
                    for cell in ws[row+row_shift]:
                        cell.value = None
                    for ci, v in enumerate(csv_values):
                        c1 = ws.cell(row=row+row_shift, column=ci + 1)
                        if em_util.isint(v):
                            c1.value = int(v)
                        elif em_util.isfloat(v):
                            c1.value = float(v)
                        else:
                            c1.value = v
                    row = row + 1


def get_diff_unified(cached: str) -> List:
    cmd1 = f"git diff {cached} --unified=0 --name-only -- *.csv"
    output_names = subprocess.run(cmd1, capture_output=True, encoding="utf-8", errors='replace').stdout
    if output_names == "":
        return []

    files = filter(lambda x: x != "", output_names.split("\n"))
    patches = []
    for fn in files:
        cmd2 = f"git diff {cached} --unified=0 {fn}"
        output_str = subprocess.run(cmd2, capture_output=True, encoding="cp932", errors='replace').stdout
        ary = output_str.split("\n")
        ary.insert(0, fn)
        patches.append(ary)
    return patches


def main():
    args = sys.argv
    if len(args) < 2:
        print(f"invalid argument")
        print(f"usage: {os.path.basename(args[0])} filepath ['staged']")
        sys.exit()

    excel_filepath = ""
    staged = ""
    if len(args) == 2:
        excel_filepath = args[1]
    if len(args) == 3 and args[2] == "staged":
        excel_filepath = args[1]
        staged = "--cached"

    workbook = openpyxl.load_workbook(excel_filepath, keep_vba=True)
    patches = get_diff_unified(staged)
    if len(patches) == 0:
        print("Not find diff, exit")
        sys.exit()

    for pt in patches:
        sheet_merge(workbook, parse_patch(pt))

    dir_name = os.path.dirname(excel_filepath)
    filename = os.path.splitext(os.path.basename(excel_filepath))[0]
    _, ext = os.path.splitext(excel_filepath)
    merged_filepath = os.path.join(dir_name, f"{filename}-merged{ext}")
    workbook.save(merged_filepath)
    print(f"save merged file: {merged_filepath}")


if __name__ == "__main__":
    main()

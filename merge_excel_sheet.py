# coding=utf-8
import os
import shutil

import csv
import subprocess
from typing import List

import openpyxl

import util


class MergeCommand:
    cmd: str
    row_index: int
    num: int
    csv_list: List[str]

    def __init__(self, cmd, row_index, num, csv_list=None):
        self.cmd = cmd
        self.row_index = row_index
        self.num = num
        self.csv_list = csv_list


class MergeData:
    sheet_name: str
    cmds: List[List[MergeCommand]]

    def __init__(self, sheet_name):
        self.sheet_name = sheet_name
        self.cmds = []


class MergeExcelSheet:
    __wb: openpyxl.Workbook
    __excel_filepath: str
    __merged_filepath: str
    __staged: str

    def __init__(self, excel_filepath: str, staged: bool):
        self.__excel_filepath = excel_filepath
        self.__staged = ""
        if staged:
            self.__staged = "--cached"
        dir_name = os.path.dirname(self.__excel_filepath)
        filename = os.path.splitext(os.path.basename(self.__excel_filepath))[0]
        _, ext = os.path.splitext(self.__excel_filepath)
        self.__merged_filepath = os.path.join(dir_name, f"{filename}-merged{ext}")

    def merge(self):
        shutil.copyfile(self.__excel_filepath, self.__merged_filepath)
        self.__wb = openpyxl.load_workbook(
            self.__merged_filepath, keep_vba=True, data_only=True)
        diff_files = self.__get_diff_files()
        patches = self.__get_patches(diff_files)
        for sheet_name, patch in patches:
            merge_data = self.__parse_patch(sheet_name, patch)
            self.__merge(self.__wb, merge_data)

    def save_merged_file(self):
        self.__wb.save(self.__merged_filepath)

    def __clear_auto_filter(self, ws):
        for i in range(ws.max_row):
            ws.row_dimensions[i].hidden = False

    def __merge(self, wb: openpyxl.Workbook, merge_data: MergeData):
        if len(merge_data.cmds) > 0:
            print(f"merge: {merge_data.sheet_name}")

        sheet_name = merge_data.sheet_name
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            # self.__clear_auto_filter(ws)
        else:
            ws = self.__wb.create_sheet(index=0, title=sheet_name)

        row_offset = util.get_row_offset(ws)
        for merge_cmds in merge_data.cmds:
            do_addrow = False
            for merge_cmd in merge_cmds:
                num = merge_cmd.num
                row_index = merge_cmd.row_index + row_offset
                if merge_cmd.cmd == "add_row":
                    print(f"\t{merge_cmd.cmd} row_index={row_index + 1}, range={num}")
                    for _ in range(num):
                        do_addrow = True
                        ws.insert_rows(row_index + 1)
                if merge_cmd.cmd == "del_row":
                    print(f"\t{merge_cmd.cmd} row_index={row_index + 1}, range={num}")
                    for ri in reversed(range(num)):
                        ws.delete_rows(row_index + ri)
                if merge_cmd.cmd == "set_row":
                    row_shift = 0
                    if do_addrow:
                        row_shift = 1
                    row = row_index
                    csv_lines = list(csv.reader(
                        merge_cmd.csv_list, quotechar='"', delimiter=',',
                        quoting=csv.QUOTE_MINIMAL, skipinitialspace=True))
                    print(f"\t{merge_cmd.cmd} row_index={row + row_shift}, range={len(csv_lines)}")
                    for csv_values in csv_lines:
                        for cell in ws[row + row_shift]:
                            cell.value = None
                        for ci, v in enumerate(csv_values):
                            c1 = ws.cell(row=row + row_shift, column=ci + 1)
                            if util.isint(v):
                                c1.value = int(v)
                            elif util.isfloat(v):
                                c1.value = float(v)
                            else:
                                c1.value = v
                        row = row + 1

    def __parse_patch(self, sheet_name: str, patches: List[str]):
        merge_data = MergeData(sheet_name)
        cmds = []
        for i, d in enumerate(patches):
            if d.startswith("@@ "):
                tmp_cmds = []
                info = d.replace("@@ ", "").replace(" @@", "")
                row_index = self.__parse_line_info(info)
                ddm = []
                ddp = []
                for x in range(i + 1, len(patches)):
                    if patches[x].startswith("-"):
                        ddm.append(patches[x])
                    elif patches[x].startswith("+"):
                        ddp.append(patches[x][1:])
                    elif patches[x].startswith("@@ "):
                        break
                    else:
                        break
                df = len(ddp) - len(ddm)
                if df > 0:
                    tmp_cmds.append(
                        MergeCommand("add_row", row_index, df))
                elif df < 0:
                    tmp_cmds.append(
                        MergeCommand("del_row", row_index, abs(df)))
                if len(ddp) > 0:
                    tmp_cmds.append(
                        MergeCommand("set_row", row_index, len(ddp), ddp))
                cmds.insert(0, tmp_cmds)
        merge_data.cmds = cmds
        return merge_data

    @staticmethod
    def __parse_line_info(line_info: str):
        diff = line_info.replace("-", "").replace("+", "").split(" ")
        pre = diff[0]
        if pre.find(",") == -1:
            return int(pre)
        else:
            return int(pre.split(",")[0])

    def __get_patches(self, files: List[str]):
        patches = []
        for fn in files:
            cmd = f"git diff {self.__staged} --unified=0 {fn}"
            output_str = subprocess.run(cmd, capture_output=True, shell=True,
                                        encoding="cp932", errors='replace').stdout
            ary: List[str] = list(filter(lambda x: x != "", output_str.split("\n")))
            for i, e in enumerate(ary):
                if e.startswith("@@ "):
                    ary = ary[i:]
                    break
            sheet_name = fn.split("/")[-1].replace(".csv", "")
            patches.append((sheet_name, ary))
        return patches

    def __get_diff_files(self):
        cmd = f"git diff {self.__staged} --unified=0 --name-only -- *.csv"
        output_names = subprocess.run(cmd, capture_output=True, shell=True,
                                      encoding="utf-8", errors='replace').stdout
        if output_names == "":
            return []

        files: List[str] = list(filter(lambda x: x != "", output_names.split("\n")))
        return files
